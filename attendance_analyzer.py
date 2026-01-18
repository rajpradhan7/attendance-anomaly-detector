from openpyxl import load_workbook

def attendance_risk(total, present, late):
    attendance_percent = (present / total) * 100

    if attendance_percent < 75:
        return "High Risk"
    elif late > 5:
        return "Medium Risk"
    else:
        return "Low Risk"


workbook = load_workbook("attendance.xlsx")
sheet = workbook.active

with open("attendance_report.txt", "w") as report:
    report.write("Attendance Risk Report\n\n")

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, total, present, late = row
        risk = attendance_risk(total, present, late)
        report.write(f"{name}: {risk}\n")

print("Attendance analysis completed. Check attendance_report.txt")
